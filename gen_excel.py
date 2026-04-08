import re

# 读取 shop_map.html 中的门店数据
with open('/Users/apple/WorkBuddy/20260405094656/图片提取/shop_map.html', 'r', encoding='utf-8') as f:
    html = f.read()

# 提取 stores 数组
stores_match = re.search(r'const stores = \[(.*?)\];', html, re.DOTALL)
if not stores_match:
    print("未找到门店数据")
    exit()

stores_str = stores_match.group(1)

# 解析每个门店对象
store_pattern = re.compile(r'{id:(\d+),name:"([^"]+)",addr:"([^"]+)",region:"([^"]+)",status:"[^"]+",lat:[\d.]+,lng:[\d.]+,district:"[^"]+",final_street:"([^"]+)"},?')
stores = []

for match in store_pattern.finditer(stores_str):
    stores.append({
        'id': match.group(1),
        'name': match.group(2),
        'addr': match.group(3),
        'region': match.group(4),
        'final_street': match.group(5)
    })

print(f"共解析 {len(stores)} 家门店")

# 提取标签数据（已面谈/已下单/未面谈）
tag_match = re.search(r'const storeTags = \{([^}]+)\}', html)
store_tags = {}
if tag_match:
    tag_str = tag_match.group(1)
    for m in re.finditer(r'(\d+):"([^"]+)"', tag_str):
        store_tags[m.group(1)] = m.group(2)

print(f"标签数据: {store_tags}")

# 生成 Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "门店数据"

# 表头
headers = ['门店编号', '门店名称', '门店地址', '门店统计标签', '门店所属街道']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入数据
for row, store in enumerate(stores, 2):
    tag = store_tags.get(store['id'], '未面谈')
    ws.cell(row=row, column=1, value=store['id'])
    ws.cell(row=row, column=2, value=store['name'])
    ws.cell(row=row, column=3, value=store['addr'])
    ws.cell(row=row, column=4, value=tag)
    ws.cell(row=row, column=5, value=store['region'])

# 统计行
ws.cell(row=len(stores)+2, column=1, value='统计：')
ws.cell(row=len(stores)+2, column=1).font = Font(bold=True)

counts = {'已面谈': 0, '已下单': 0, '未面谈': 0}
for tag in store_tags.values():
    if tag in counts:
        counts[tag] += 1

yimiantan = counts['已面谈']
yixiadan = counts['已下单']
weimiantan = len(stores) - yimiantan - yixiadan

ws.cell(row=len(stores)+3, column=1, value=f'已面谈: {yimiantan}')
ws.cell(row=len(stores)+4, column=1, value=f'已下单: {yixiadan}')
ws.cell(row=len(stores)+5, column=1, value=f'未面谈: {weimiantan}')

# 列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12

# 保存
output_path = '/Users/apple/WorkBuddy/20260405094656/图片提取/门店数据.xlsx'
wb.save(output_path)
print(f"Excel 已生成: {output_path}")